#!/usr/bin/env python
# -*- coding: utf-8 -*-

from tkinter import * 
from tkinter import ttk

from openpyxl import load_workbook
import urllib.request
import json
from datetime import datetime,date
import tkinter.font as tkFont
# Genera Excel
class Aplicacion():
    def __init__(self):
        self.raiz = Tk()
        #define tamaño de la ventana
        self.raiz.geometry('700x600')
        self.raiz.title("FORMULARIO DE VINCULACIÓN/ACTUALIZACION V1.1 COOP. SAN MARTÍN")
        # Declara variables de control
        self.nsocio = StringVar()
        self.mostrar_respuesta = StringVar()

        # Define trazas con variables de control de los widgets Entry()
        # para detectar cambios en los datos. Si se producen cambios
        # se llama a la función 'self.calcular' para validación y para
        # calcular importe a pagar

        #estilo de fuente
        fontStyle = tkFont.Font(family="Arial", size=14)
        #Estilos para ttk
        s = ttk.Style()
        s.configure('my.TButton', font=('Arial', 14))

        self.mostrar_respuesta.trace('w', self.buscar)

        self.boton2 = ttk.Button(self.raiz, text="Buscar", command=self.buscar,style='my.TButton')

        self.etiq1 = ttk.Label(self.raiz, text="INGRESE EL NÚMERO DE SOCIO:", font=fontStyle)
        self.etiq2 = ttk.Entry(self.raiz, textvariable=self.nsocio, width=10,font=fontStyle)

        self.etiq5 = ttk.Label(self.raiz, text="DATOS  DEL SOCIO:",font=fontStyle)
        self.etiq6 = ttk.Label(self.raiz, textvariable=self.mostrar_respuesta, foreground="yellow", background="black", borderwidth=0, font=fontStyle)
        self.separ1 = ttk.Separator(self.raiz, orient=HORIZONTAL)

        self.boton1 = ttk.Button(self.raiz, text="GENERAR EXCEL", command=self.calcular, style='my.TButton')
        self.etiq1.pack(side=TOP, fill=BOTH, expand=True,padx=2, pady=2)
        self.etiq2.pack(side=TOP, fill=BOTH, expand=True,padx=2, pady=2)
        self.boton2.pack(side=TOP, fill=BOTH, expand=True,padx=250, pady=2)
        self.etiq5.pack(side=TOP, fill=BOTH, expand=True, padx=2, pady=2)
        self.etiq6.pack(side=TOP, fill=BOTH, expand=True, padx=2, pady=2)
        self.separ1.pack(side=TOP, fill=BOTH, expand=True, padx=2, pady=2)
        self.boton1.pack(side=TOP, fill=BOTH, expand=True, padx=250, pady=2)
        self.raiz.mainloop()

    def buscar(self, *args):
         #produccion
         #with urllib.request.urlopen("http://10.1.0.173:8080/integrado/plataforma/socio/"+self.etiq2.get()) as url:
         #   data = json.loads(url.read().decode())

         with urllib.request.urlopen("http://10.1.0.173:8080/integrado/plataforma/socio/"+self.etiq2.get()) as url:
            data = json.loads(url.read().decode())
            datos_socio="NRO DE SOCIO:"+data[0]['gbagecage']+"\n"+"NOMBRE:"+data[0]['nombre']+"\n"+"CI:"+data[0]['ci']
            self.mostrar_respuesta.set(datos_socio)
    def calcular(self, *args):
        # Función para validar datos y calcular importe a pagar
        # My code   
        #declaracion de variables
        fecha_actual=date.today()    
        #with urllib.request.urlopen("http://localhost:8080/plataforma/socio/"+self.etiq2.get()) as url:
        #production endpoints
        #with urllib.request.urlopen("http://10.1.0.173:8080/integrado/plataforma/socio/"+self.etiq2.get()) as url:
        #   data = json.loads(url.read().decode())

        #with urllib.request.urlopen("http://10.1.0.173:8080/integrado/declaracion/sel/"+self.etiq2.get()) as url:
        #   declaracion = json.loads(url.read().decode())
        
        with urllib.request.urlopen("http://10.1.0.173:8080/integrado/plataforma/socio/"+self.etiq2.get()) as url:
           data = json.loads(url.read().decode())

        with urllib.request.urlopen("http://10.1.0.173:8080/integrado/declaracion/sel/"+self.etiq2.get()) as url:
           declaracion = json.loads(url.read().decode())
        with urllib.request.urlopen("http://10.1.0.173:8080/integrado/declaracion/activos/"+self.etiq2.get()) as url:
           activos = json.loads(url.read().decode())
        with urllib.request.urlopen("http://10.1.0.173:8080/integrado/declaracion/ingresosfijos/"+self.etiq2.get()) as url:
           ingresosfijos = json.loads(url.read().decode())

        with urllib.request.urlopen("http://10.1.0.173:8080/integrado/declaracion/pasivos/"+self.etiq2.get()) as url:
           pasivos = json.loads(url.read().decode())

        with urllib.request.urlopen("http://10.1.0.173:8080/integrado/declaracion/gastosfijos/"+self.etiq2.get()) as url:
           gastosfijos = json.loads(url.read().decode())

        #localendpoints
        # print(data[0]['nombre'])
        print("NRO DE SOCIO:"+data[0]['gbagecage']+"\n" +
          "NOMBRE:"+data[0]['nombre']+"\n"+"CI:"+data[0]['ci'])
        print(len(activos))

        # declaracion de variables:
        # importamos load_workbook
        # ruta de nuestro archivo
        filesheet = "plantilla.xlsx"
        # creamos el objeto load_workbook
        wb = load_workbook(filesheet)
        # Seleccionamos el archivo
        sheet = wb.active
        # parametros
        # Ingresamos el valor nombre en la celda 'A12'
        #año         
        sheet['AG1'] = fecha_actual
        sheet['A12'] = data[0]['nombre']
        sheet['AE12'] = data[0]['ci']
        sheet['A16'] = data[0]['estado_civil']
        sheet['U16'] = data[0]['direccion']
        sheet['k16'] = data[0]['nacionalidad']
        sheet['A19'] = data[0]['profesion']
        sheet['I28'] = data[0]['fecha_nacimiento']
        sheet['AL28'] = data[0]['celular']
        sheet['R19'] = data[0]['cargo']
        #sheet['AB19'] = data[0]['antiguedad']
        sheet['K17'] = data[0]['tipo_vivienda']
        sheet['K19'] = data[0]['nit']
        sheet['AE59'] = data[0]['gbdaccand']
        sheet['F59'] = data[0]['gbdacmail']
        sheet['G88'] = data[0]['gbdacrefp']
        sheet['K26'] = data[0]['gbdacrefo']
        #activos
        if len(activos)!=0:
            for i in range (0,len(activos)):
                sheet[f'A{43+i}']=activos[i]['concepto']
                sheet[f'P{43+i}']=activos[i]['valor']
        #ingresos
        if len(pasivos)!=0:
            for i in range (0,len(pasivos)):
                sheet[f'A{54+i}']=pasivos[i]['concepto']
                sheet[f'P{54+i}']=pasivos[i]['valor']
        if len(ingresosfijos)!=0:
            for i in range (0,len(ingresosfijos)):
                sheet[f'W{43+i}']=ingresosfijos[i]['concepto']
                sheet[f'AN{43+i}']=ingresosfijos[i]['valor']
        if len(gastosfijos)!=0:
            for i in range (0,len(gastosfijos)):
                sheet[f'W{50+i}']=gastosfijos[i]['concepto']
                sheet[f'AN{50+i}']=gastosfijos[i]['valor']
 
        #totales
        if len(declaracion)!=0:
            sheet['P52'] = declaracion[0]['t_activos']
            sheet['P57'] = declaracion[0]['t_pasivos']
            sheet['P58'] = declaracion[0]['t_patrimonio']
            sheet['AN48'] = declaracion[0]['t_ingresosfijos']
            sheet['AN57'] = declaracion[0]['t_gastosfijos']
        
        booleano=data[0]['antiguedad']
        
        if booleano==None:
         sheet['AH19'] = " "
        else:
         #anio = datetime.strptime(str(data[0]['antiguedad']), "%Y-%m-%d")
         sheet['AH19'] = data[0]['antiguedad']
       # sheet['AH19'] = data[0]['antiguedad']

        # Guardamos el archivo con los cambios
        wb.save(data[0]['gbagecage']+" "+data[0]['nombre']+".xlsx")
        if(wb.save):
         self.mostrar_respuesta.set("Se ha guardado los datos en el archivo Excel..")
        else:
          self.mostrar_respuesta.set("No se ha guardado los cambios")

def main():
    mi_app = Aplicacion()
    return 0

if __name__ == '__main__':
    main()
